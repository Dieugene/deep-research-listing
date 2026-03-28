"""
Generate data appendix: Excel workbook + LaTeX document
for institutional clustering analysis.
"""

import csv
import math
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# -- paths ------------------------------------------------------------------
BASE = Path(r"D:\_workspace\deep-research-listing\03_data\institutional")
MASTER = BASE / "master_factors.csv"
TRAJECTORY = BASE / "trajectory_panel.csv"
CLUSTER_S1 = BASE / "cluster_assignments.csv"
CLUSTER_S4 = BASE / "stage4_cluster_assignments.csv"
LEGAL_ORIGINS = BASE / "legal_origins_laporta.xlsx"
WGI_DETAILED = BASE / "F4-6 P_Data_Extract_From_Worldwide_Governance_Indicators_detailed.xlsx"
OUT_XLSX = BASE / "source_data_appendix.xlsx"
OUT_TEX = BASE / "appendix_source_data.tex"

# -- country name mapping (WGI -> our names) --------------------------------
COUNTRY_MAP = {
    "Korea, Rep.": "South Korea",
    "Hong Kong SAR, China": "Hong Kong",
    "Hong Kong, China": "Hong Kong",
    "Turkiye": "Turkey",
    "T\u00fcrkiye": "Turkey",
    "Czech Republic": "Czech Republic",
    "Czechia": "Czech Republic",
    "Russian Federation": "Russia",
    "Egypt, Arab Rep.": "Egypt",
    "Taiwan, China": "Taiwan",
}

YEARS = [str(y) for y in range(2009, 2025)]


# -- helpers ----------------------------------------------------------------
def load_csv(path):
    df = pd.read_csv(path)
    return df


def style_header(ws, ncols, row=1):
    bold = Font(bold=True)
    fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        length = min_width
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def freeze_top(ws, row=2):
    ws.freeze_panes = f"A{row}"


def add_source_row(ws, text, ncols):
    """Insert a merged source-note row at the top and push data down."""
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = Font(italic=True, size=9)
    cell.alignment = Alignment(wrap_text=True)


# -- WGI detailed data loader -----------------------------------------------
def load_wgi_component(series_code: str) -> pd.DataFrame:
    """Load one WGI component from the detailed Excel, filtered to 48 jurisdictions.
    Returns DataFrame with index=country (our naming), columns=2009..2024."""
    raw = pd.read_excel(WGI_DETAILED, sheet_name="Data")
    sub = raw[raw["Series Code"] == series_code].copy()
    sub["country"] = sub["Country Name"].map(COUNTRY_MAP).fillna(sub["Country Name"])
    master_countries = set(load_csv(MASTER)["country"].unique())
    sub = sub[sub["country"].isin(master_countries)].copy()
    sub = sub.set_index("country")
    year_cols = {f"{y} [YR{y}]": str(y) for y in range(2009, 2025)}
    result = sub[list(year_cols.keys())].rename(columns=year_cols)
    result = result.apply(pd.to_numeric, errors="coerce")
    result = result.sort_index()
    return result


# -- LaTeX helpers -----------------------------------------------------------
def tex_escape(s):
    """Escape special LaTeX characters."""
    if not isinstance(s, str):
        s = str(s)
    replacements = {
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s


def fmt(val, decimals=3):
    if pd.isna(val) or val == "" or val is None:
        return "---"
    try:
        v = float(val)
        return f"{v:.{decimals}f}"
    except (ValueError, TypeError):
        return str(val)


def fmt_int(val):
    if pd.isna(val) or val == "" or val is None:
        return "---"
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return str(val)


# =====================================================================
#  PART 1: EXCEL WORKBOOK
# =====================================================================
def build_excel():
    wb = openpyxl.Workbook()
    master = load_csv(MASTER)
    traj = load_csv(TRAJECTORY)
    cl1 = load_csv(CLUSTER_S1)
    cl4 = load_csv(CLUSTER_S4)
    lo = pd.read_excel(LEGAL_ORIGINS)

    # Load WGI component-level annual data
    f4_annual = load_wgi_component("GOV_WGI_RQ.EST")
    f5_annual = load_wgi_component("GOV_WGI_RL.EST")
    f6_annual = load_wgi_component("GOV_WGI_PV.EST")

    # -- Sheet 1: Sources ---------------------------------------------------
    ws1 = wb.active
    ws1.title = "Источники данных"
    headers = ["Код", "Название", "Источник", "URL", "Дата данных", "Лицензия"]
    rows = [
        ["F1", "Правовая семья (Legal Origin)",
         'La Porta, Lopez-de-Silanes, Shleifer (2008). "The Economic Consequences of Legal Origins", Journal of Economic Literature',
         "https://faculty.tuck.dartmouth.edu/rafael-laporta/research-publications/",
         "2008", "Academic"],
        ["F1x", "Anti-Self-Dealing Index (ASDI)",
         "Djankov et al. (2008), via La Porta et al. dataset",
         "https://faculty.tuck.dartmouth.edu/images/uploads/faculty/rafael-laporta/EconomicCon_data.xls",
         "2003", "Academic"],
        ["F2a", "Extent of Disclosure Index",
         "World Bank, Doing Business 2020 (Protecting Minority Investors)",
         "https://archive.doingbusiness.org/en/data/exploretopics/protecting-minority-investors",
         "01.05.2019", "CC BY-4.0"],
        ["F2b", "Extent of Director Liability Index",
         "World Bank, Doing Business 2020 (Protecting Minority Investors)",
         "https://archive.doingbusiness.org/en/data/exploretopics/protecting-minority-investors",
         "01.05.2019", "CC BY-4.0"],
        ["F2c", "Ease of Shareholder Suits Index",
         "World Bank, Doing Business 2020 (Protecting Minority Investors)",
         "https://archive.doingbusiness.org/en/data/exploretopics/protecting-minority-investors",
         "01.05.2019", "CC BY-4.0"],
        ["F4", "Regulatory Quality (WGI)",
         "World Bank, Worldwide Governance Indicators, 2025 Revision",
         "https://www.govindicators.org",
         "2009-2024 (годовые)", "CC BY-4.0"],
        ["F5", "Rule of Law (WGI)",
         "World Bank, Worldwide Governance Indicators, 2025 Revision",
         "https://www.govindicators.org",
         "2009-2024 (годовые)", "CC BY-4.0"],
        ["F6", "Political Stability (WGI)",
         "World Bank, Worldwide Governance Indicators, 2025 Revision",
         "https://www.govindicators.org",
         "2009-2024 (годовые)", "CC BY-4.0"],
        ["F7", "Market Capitalization / GDP (%)",
         "World Bank, World Development Indicators (WDI) via WFE",
         "https://databank.worldbank.org/reports.aspx?source=2&series=CM.MKT.LCAP.GD.ZS&country=#",
         "Последний доступный год", "CC BY-4.0"],
        ["Fx", "Gross Domestic Savings / GDP (%)",
         "World Bank, World Development Indicators (WDI)",
         "https://databank.worldbank.org/reports.aspx?source=2&series=CM.MKT.LCAP.GD.ZS&country=#",
         "Последний доступный год", "CC BY-4.0"],
    ]
    ws1.append(headers)
    for r in rows:
        ws1.append(r)
    style_header(ws1, len(headers))
    auto_width(ws1)
    freeze_top(ws1)

    # -- Sheet 2: F1_Legal_Origin_ASDI --------------------------------------
    ws2 = wb.create_sheet("F1_Legal_Origin_ASDI")
    lo_headers = list(lo.columns)
    ws2.append(lo_headers)
    for _, row in lo.iterrows():
        ws2.append([row[c] for c in lo_headers])
    style_header(ws2, len(lo_headers))
    add_source_row(ws2, "Источник: La Porta et al. (2008). The Economic Consequences of Legal Origins", len(lo_headers))
    auto_width(ws2)
    freeze_top(ws2, row=3)

    # -- Sheet 3: F2_Investor_Protection ------------------------------------
    ws3 = wb.create_sheet("F2_Investor_Protection")
    cols3 = ["country", "F2a_disclosure_val", "F2a_disclosure_year",
             "F2b_director_liability_val", "F2b_director_liability_year",
             "F2c_shareholder_suits_val", "F2c_shareholder_suits_year",
             "F2_composite_val"]
    nice3 = ["Country", "F2a Disclosure", "F2a Year",
             "F2b Dir.Liability", "F2b Year",
             "F2c Shareh.Suits", "F2c Year",
             "F2 Composite"]
    ws3.append(nice3)
    for _, row in master.sort_values("country").iterrows():
        vals = []
        for c in cols3:
            v = row[c]
            if "year" in c:
                vals.append(int(v) if pd.notna(v) else "")
            elif c == "F2_composite_val":
                vals.append(round(v, 2) if pd.notna(v) else "")
            elif "val" in c:
                vals.append(int(v) if pd.notna(v) else "")
            else:
                vals.append(v)
        ws3.append(vals)
    style_header(ws3, len(nice3))
    add_source_row(ws3, "Источник: World Bank, Doing Business 2020 --- Protecting Minority Investors", len(nice3))
    auto_width(ws3)
    freeze_top(ws3, row=3)

    # -- Sheet 4: F4-F6_WGI_2024 -------------------------------------------
    ws4 = wb.create_sheet("F4-F6_WGI_2024")
    nice4 = ["Country", "F4 Reg.Quality 2024", "F5 Rule of Law 2024",
             "F6 Pol.Stability 2024", "WGI Composite 2024"]
    ws4.append(nice4)
    for _, row in master.sort_values("country").iterrows():
        ws4.append([
            row["country"],
            round(row["F4_reg_quality_2024"], 3) if pd.notna(row["F4_reg_quality_2024"]) else "",
            round(row["F5_rule_of_law_2024"], 3) if pd.notna(row["F5_rule_of_law_2024"]) else "",
            round(row["F6_pol_stability_2024"], 3) if pd.notna(row["F6_pol_stability_2024"]) else "",
            round(row["WGI_composite_2024"], 3) if pd.notna(row["WGI_composite_2024"]) else "",
        ])
    style_header(ws4, len(nice4))
    add_source_row(ws4, "Источник: World Bank, Worldwide Governance Indicators, 2025 Revision (https://www.govindicators.org)", len(nice4))
    auto_width(ws4)
    freeze_top(ws4, row=3)

    # -- Sheet 5: F4_Regulatory_Quality_Annual ------------------------------
    def write_annual_sheet(wb, df, sheet_name, source_note):
        ws = wb.create_sheet(sheet_name)
        hdrs = ["Country"] + YEARS
        ws.append(hdrs)
        for country in df.index:
            vals = [country]
            for y in YEARS:
                v = df.loc[country, y] if y in df.columns else ""
                vals.append(round(float(v), 3) if pd.notna(v) else "")
            ws.append(vals)
        style_header(ws, len(hdrs))
        add_source_row(ws, source_note, len(hdrs))
        auto_width(ws, min_width=7, max_width=12)
        freeze_top(ws, row=3)

    write_annual_sheet(wb, f4_annual, "F4_Regulatory_Quality_Annual",
                       "F4 Regulatory Quality (Estimate), 2009-2024. Источник: World Bank WGI 2025 Revision")
    write_annual_sheet(wb, f5_annual, "F5_Rule_of_Law_Annual",
                       "F5 Rule of Law (Estimate), 2009-2024. Источник: World Bank WGI 2025 Revision")
    write_annual_sheet(wb, f6_annual, "F6_Political_Stability_Annual",
                       "F6 Political Stability (Estimate), 2009-2024. Источник: World Bank WGI 2025 Revision")

    # -- Sheet 8: F4-F6_WGI_Composite_Annual (renamed from F4-F6_WGI_Annual)
    ws_comp = wb.create_sheet("F4-F6_WGI_Composite_Annual")
    traj_headers = ["Country"] + YEARS
    ws_comp.append(traj_headers)
    traj_sorted = traj.sort_values(traj.columns[0])
    for _, row in traj_sorted.iterrows():
        vals = [row.iloc[0]]  # country
        for y in YEARS:
            v = row[y] if y in row.index else ""
            vals.append(round(float(v), 3) if pd.notna(v) else "")
        ws_comp.append(vals)
    style_header(ws_comp, len(traj_headers))
    add_source_row(ws_comp, "WGI Composite (среднее F4+F5+F6), 2009-2024. Источник: World Bank WGI 2025 Revision", len(traj_headers))
    auto_width(ws_comp, min_width=7, max_width=12)
    freeze_top(ws_comp, row=3)

    # -- Sheet 9: F7_Market_Structure ---------------------------------------
    ws6 = wb.create_sheet("F7_Market_Structure")
    cols6_nice = ["Country", "MktCap/GDP (%)", "MktCap Year", "Savings/GDP (%)", "Savings Year"]
    ws6.append(cols6_nice)
    for _, row in master.sort_values("country").iterrows():
        ws6.append([
            row["country"],
            round(row["F7_mktcap_gdp_val"], 2) if pd.notna(row["F7_mktcap_gdp_val"]) else "",
            int(row["F7_mktcap_gdp_year"]) if pd.notna(row["F7_mktcap_gdp_year"]) else "",
            round(row["Fx_savings_gdp_val"], 2) if pd.notna(row["Fx_savings_gdp_val"]) else "",
            int(row["Fx_savings_gdp_year"]) if pd.notna(row["Fx_savings_gdp_year"]) else "",
        ])
    style_header(ws6, len(cols6_nice))
    add_source_row(ws6, "Источник: World Bank, World Development Indicators (WDI)", len(cols6_nice))
    auto_width(ws6)
    freeze_top(ws6, row=3)

    # -- Sheet 10: Clustering Stage I ----------------------------------------
    ws7 = wb.create_sheet("Кластеризация_Этап_I")
    cl1_cols = ["country", "cluster_A", "legal_origin", "market_group", "WGI_trajectory"]
    cl1_nice = ["Country", "Cluster (Variant A)", "Legal Origin", "Market Group", "WGI Trajectory"]
    ws7.append(cl1_nice)
    for _, row in cl1.sort_values("country").iterrows():
        ws7.append([row[c] for c in cl1_cols])
    style_header(ws7, len(cl1_nice))
    add_source_row(ws7, "Результаты кластеризации Этап I (K-Prototypes, Variant A), 43 юрисдикции", len(cl1_nice))
    auto_width(ws7)
    freeze_top(ws7, row=3)

    # -- Sheet 11: Clustering Stage IV ---------------------------------------
    ws8 = wb.create_sheet("Кластеризация_Этап_IV")
    cl4_cols = ["entity", "cluster_mfa", "silhouette_mfa"]
    cl4_nice = ["Entity", "Cluster (MFA)", "Silhouette (MFA)"]
    ws8.append(cl4_nice)
    for _, row in cl4.sort_values("entity").iterrows():
        ws8.append([
            row["entity"],
            int(row["cluster_mfa"]),
            round(row["silhouette_mfa"], 4),
        ])
    style_header(ws8, len(cl4_nice))
    add_source_row(ws8, "Результаты кластеризации Этап IV (MFA + K-Means), 49 объектов", len(cl4_nice))
    auto_width(ws8)
    freeze_top(ws8, row=3)

    wb.save(OUT_XLSX)
    print(f"Excel saved: {OUT_XLSX}")


# =====================================================================
#  PART 2: LATEX DOCUMENT
# =====================================================================
def build_latex():
    master = load_csv(MASTER).sort_values("country")
    traj = load_csv(TRAJECTORY)
    traj = traj.sort_values(traj.columns[0])
    cl1 = load_csv(CLUSTER_S1).sort_values("country")
    cl4 = load_csv(CLUSTER_S4).sort_values("entity")
    lo = pd.read_excel(LEGAL_ORIGINS).sort_values("Country")

    # Load WGI component-level annual data
    f4_annual = load_wgi_component("GOV_WGI_RQ.EST")
    f5_annual = load_wgi_component("GOV_WGI_RL.EST")
    f6_annual = load_wgi_component("GOV_WGI_PV.EST")

    lines = []
    L = lines.append

    # -- preamble -----------------------------------------------------------
    L(r"\documentclass[a4paper,11pt]{article}")
    L(r"\usepackage[T2A]{fontenc}")
    L(r"\usepackage[utf8]{inputenc}")
    L(r"\usepackage[russian,english]{babel}")
    L(r"\usepackage[margin=2cm]{geometry}")
    L(r"\usepackage{booktabs}")
    L(r"\usepackage{longtable}")
    L(r"\usepackage{array}")
    L(r"\usepackage{pdflscape}")
    L(r"\usepackage[hyphens,spaces,obeyspaces]{url}")
    L(r"\usepackage{hyperref}")
    L(r"\usepackage{xcolor}")
    L(r"\hypersetup{colorlinks=true,linkcolor=blue,urlcolor=blue,citecolor=blue}")
    L(r"\expandafter\def\expandafter\UrlBreaks\expandafter{\UrlBreaks"
      r"\do\a\do\b\do\c\do\d\do\e\do\f\do\g\do\h\do\i\do\j\do\k\do\l"
      r"\do\m\do\n\do\o\do\p\do\q\do\r\do\s\do\t\do\u\do\v\do\w\do\x"
      r"\do\y\do\z\do\A\do\B\do\C\do\D\do\E\do\F\do\G\do\H\do\I\do\J"
      r"\do\K\do\L\do\M\do\N\do\O\do\P\do\Q\do\R\do\S\do\T\do\U\do\V"
      r"\do\W\do\X\do\Y\do\Z\do\0\do\1\do\2\do\3\do\4\do\5\do\6\do\7"
      r"\do\8\do\9\do\/\do\.\do\-\do\_\do\~\do\=\do\&\do\?\do\#}")
    L(r"")
    L(r"\title{Приложение: Исходные данные для кластеризации юрисдикций\\по институциональным факторам}")
    L(r"\author{}")
    L(r"\date{24 марта 2026 г.}")
    L(r"")
    L(r"\begin{document}")
    L(r"\maketitle")
    L(r"\tableofcontents")
    L(r"\newpage")

    # -- Section 1: Sources -------------------------------------------------
    L(r"")
    L(r"\section{Источники данных}")
    L(r"")
    L(r"В настоящем приложении представлены исходные данные, использованные для кластеризации 48 юрисдикций")
    L(r"по институциональным факторам листинга ценных бумаг. Ниже перечислены источники данных.")
    L(r"")
    L(r"\begin{longtable}{p{1cm}p{3.5cm}p{5.5cm}p{2cm}p{2cm}}")
    L(r"\toprule")
    L(r"\textbf{Код} & \textbf{Название} & \textbf{Источник} & \textbf{Дата} & \textbf{Лицензия} \\")
    L(r"\midrule")
    L(r"\endhead")

    sources = [
        ("F1", "Правовая семья (Legal Origin)",
         r"La Porta, Lopez-de-Silanes, Shleifer (2008). \textit{The Economic Consequences of Legal Origins}, Journal of Economic Literature",
         "2008", "Academic"),
        ("F1x", "Anti-Self-Dealing Index",
         r"Djankov et al. (2008), via La Porta et al.",
         "2003", "Academic"),
        ("F2a", "Extent of Disclosure Index",
         r"World Bank, Doing Business 2020",
         "05.2019", "CC BY-4.0"),
        ("F2b", "Dir. Liability Index",
         r"World Bank, Doing Business 2020",
         "05.2019", "CC BY-4.0"),
        ("F2c", "Shareholder Suits Index",
         r"World Bank, Doing Business 2020",
         "05.2019", "CC BY-4.0"),
        ("F4", "Regulatory Quality",
         r"World Bank, WGI 2025 Revision",
         "2009--2024", "CC BY-4.0"),
        ("F5", "Rule of Law",
         r"World Bank, WGI 2025 Revision",
         "2009--2024", "CC BY-4.0"),
        ("F6", "Political Stability",
         r"World Bank, WGI 2025 Revision",
         "2009--2024", "CC BY-4.0"),
        ("F7", r"Market Cap / GDP (\%)",
         r"World Bank, WDI via WFE",
         "посл. год", "CC BY-4.0"),
        ("Fx", r"Savings / GDP (\%)",
         r"World Bank, WDI",
         "посл. год", "CC BY-4.0"),
    ]
    for code, name, src, date, lic in sources:
        L(f"{code} & {name} & {src} & {date} & {lic} \\\\")
    L(r"\bottomrule")
    L(r"\end{longtable}")
    L(r"")
    L(r"\noindent Полные URL источников:")
    L(r"\begin{itemize}")
    L(r"  \item F1, F1x: \url{https://faculty.tuck.dartmouth.edu/rafael-laporta/research-publications/}")
    L(r"  \item F2a--F2c: \url{https://archive.doingbusiness.org/en/data/exploretopics/protecting-minority-investors}")
    L(r"  \item F4--F6: \url{https://www.govindicators.org}")
    L(r"  \item F7, Fx: \url{https://databank.worldbank.org/reports.aspx?source=2&series=CM.MKT.LCAP.GD.ZS}")
    L(r"\end{itemize}")

    # -- Section 2: F1 ------------------------------------------------------
    L(r"")
    L(r"\section{F1 --- Правовая семья и Anti-Self-Dealing Index}")
    L(r"")
    L(r"\begin{longtable}{lll}")
    L(r"\toprule")
    L(r"\textbf{Country} & \textbf{Legal Origin} & \textbf{ASDI} \\")
    L(r"\midrule")
    L(r"\endhead")

    master_countries = set(master["country"].tolist())
    lo_filtered = lo[lo["Country"].isin(master_countries)].sort_values("Country")
    for _, row in lo_filtered.iterrows():
        country = tex_escape(row["Country"])
        origin = tex_escape(str(row["Legal Origin"]))
        asdi = fmt(row["ASDI"], 2)
        L(f"{country} & {origin} & {asdi} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{3}{l}{\footnotesize Источник: La Porta et al. (2008). Данные ASDI --- 2003 г.} \\")
    L(r"\end{longtable}")

    # -- Section 3: F2 ------------------------------------------------------
    L(r"")
    L(r"\section{F2 --- Защита миноритарных инвесторов (Doing Business 2020)}")
    L(r"")
    L(r"\begin{longtable}{lcccc}")
    L(r"\toprule")
    L(r"\textbf{Country} & \textbf{F2a Disclosure} & \textbf{F2b Dir.Liab.} & \textbf{F2c Shareh.Suits} & \textbf{F2 Composite} \\")
    L(r"\midrule")
    L(r"\endhead")
    for _, row in master.iterrows():
        country = tex_escape(row["country"])
        f2a = fmt_int(row["F2a_disclosure_val"])
        f2b = fmt_int(row["F2b_director_liability_val"])
        f2c = fmt_int(row["F2c_shareholder_suits_val"])
        f2comp = fmt(row["F2_composite_val"], 2)
        L(f"{country} & {f2a} & {f2b} & {f2c} & {f2comp} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{5}{l}{\footnotesize Источник: World Bank, Doing Business 2020. Дата данных: май 2019 г.} \\")
    L(r"\end{longtable}")

    # -- Section 4: F4-F6 WGI 2024 ------------------------------------------
    L(r"")
    L(r"\section{F4--F6 --- Worldwide Governance Indicators (2024)}")
    L(r"")
    L(r"\begin{longtable}{lcccc}")
    L(r"\toprule")
    L(r"\textbf{Country} & \textbf{F4 Reg.Quality} & \textbf{F5 Rule of Law} & \textbf{F6 Pol.Stability} & \textbf{WGI Composite} \\")
    L(r"\midrule")
    L(r"\endhead")
    for _, row in master.iterrows():
        country = tex_escape(row["country"])
        f4 = fmt(row["F4_reg_quality_2024"], 3)
        f5 = fmt(row["F5_rule_of_law_2024"], 3)
        f6 = fmt(row["F6_pol_stability_2024"], 3)
        wgi = fmt(row["WGI_composite_2024"], 3)
        L(f"{country} & {f4} & {f5} & {f6} & {wgi} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{5}{l}{\footnotesize Источник: World Bank, WGI 2025 Revision (\url{https://www.govindicators.org}).} \\")
    L(r"\end{longtable}")

    # -- Section 5: WGI Annual Panels (landscape) ----------------------------
    L(r"")
    L(r"\section{F4--F6 --- Годовые данные WGI (2009--2024)}")

    year_cols_spec = " ".join(["r"] * 16)
    header_row = r"\textbf{Country} & " + " & ".join([r"\textbf{" + y + "}" for y in YEARS]) + r" \\"

    def emit_annual_landscape_table(df, title, subtitle, footnote):
        L(r"")
        L(r"\subsection{" + title + "}")
        L(r"")
        L(r"\begin{landscape}")
        L(r"\footnotesize")
        L(r"\begin{longtable}{l" + year_cols_spec + "}")
        L(r"\toprule")
        L(header_row)
        L(r"\midrule")
        L(r"\endhead")
        for country in df.index:
            c_esc = tex_escape(country)
            vals = []
            for y in YEARS:
                v = df.loc[country, y] if y in df.columns else None
                vals.append(fmt(v, 3))
            L(f"{c_esc} & " + " & ".join(vals) + r" \\")
        L(r"\bottomrule")
        L(r"\multicolumn{17}{l}{\footnotesize " + footnote + r"} \\")
        L(r"\end{longtable}")
        L(r"\end{landscape}")

    emit_annual_landscape_table(
        f4_annual,
        r"F4 --- Regulatory Quality (2009--2024)",
        "F4 Regulatory Quality Estimate",
        r"Источник: World Bank, WGI 2025 Revision (\url{https://www.govindicators.org}).",
    )

    emit_annual_landscape_table(
        f5_annual,
        r"F5 --- Rule of Law (2009--2024)",
        "F5 Rule of Law Estimate",
        r"Источник: World Bank, WGI 2025 Revision (\url{https://www.govindicators.org}).",
    )

    emit_annual_landscape_table(
        f6_annual,
        r"F6 --- Political Stability (2009--2024)",
        "F6 Political Stability Estimate",
        r"Источник: World Bank, WGI 2025 Revision (\url{https://www.govindicators.org}).",
    )

    # 5.4 WGI Composite (from trajectory_panel.csv)
    L(r"")
    L(r"\subsection{WGI Composite (2009--2024)}")
    L(r"")
    L(r"\begin{landscape}")
    L(r"\footnotesize")
    L(r"\begin{longtable}{l" + year_cols_spec + "}")
    L(r"\toprule")
    L(header_row)
    L(r"\midrule")
    L(r"\endhead")
    for _, row in traj.iterrows():
        country = tex_escape(str(row.iloc[0]))
        vals = []
        for y in YEARS:
            v = row[y] if y in row.index else None
            vals.append(fmt(v, 3))
        L(f"{country} & " + " & ".join(vals) + r" \\")
    L(r"\bottomrule")
    L(r"\multicolumn{17}{l}{\footnotesize WGI Composite = среднее (F4 + F5 + F6). Источник: World Bank, WGI 2025 Revision.} \\")
    L(r"\end{longtable}")
    L(r"\end{landscape}")

    # -- Section 6: F7 Market Structure --------------------------------------
    L(r"")
    L(r"\section{F7 --- Рыночная структура}")
    L(r"")
    L(r"\begin{longtable}{lrclrc}")
    L(r"\toprule")
    L(r"\textbf{Country} & \textbf{MktCap/GDP (\%)} & \textbf{Год} & & \textbf{Savings/GDP (\%)} & \textbf{Год} \\")
    L(r"\midrule")
    L(r"\endhead")
    for _, row in master.iterrows():
        country = tex_escape(row["country"])
        mc = fmt(row["F7_mktcap_gdp_val"], 2)
        mc_y = fmt_int(row["F7_mktcap_gdp_year"])
        sv = fmt(row["Fx_savings_gdp_val"], 2)
        sv_y = fmt_int(row["Fx_savings_gdp_year"])
        L(f"{country} & {mc} & {mc_y} & & {sv} & {sv_y} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{6}{l}{\footnotesize Источник: World Bank, WDI (\url{https://databank.worldbank.org}).} \\")
    L(r"\end{longtable}")

    # -- Section 7: Clustering results ---------------------------------------
    L(r"")
    L(r"\section{Результаты кластеризации}")
    L(r"")
    L(r"\subsection{Этап I: K-Prototypes (43 юрисдикции)}")
    L(r"")
    L(r"\begin{longtable}{lcllc}")
    L(r"\toprule")
    L(r"\textbf{Country} & \textbf{Cluster} & \textbf{Legal Origin} & \textbf{Market} & \textbf{WGI Trajectory} \\")
    L(r"\midrule")
    L(r"\endhead")
    for _, row in cl1.iterrows():
        country = tex_escape(row["country"])
        cl = str(int(row["cluster_A"]))
        lo_val = tex_escape(str(row["legal_origin"]))
        mg = tex_escape(str(row["market_group"]))
        traj_val = tex_escape(str(row["WGI_trajectory"]))
        L(f"{country} & {cl} & {lo_val} & {mg} & {traj_val} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{5}{l}{\footnotesize Этап I: K-Prototypes, Variant A. 43 юрисдикции (без Taiwan, Denmark, Finland, Sweden, Italy).} \\")
    L(r"\end{longtable}")

    L(r"")
    L(r"\subsection{Этап IV: MFA + K-Means (49 объектов)}")
    L(r"")
    L(r"\begin{longtable}{lcc}")
    L(r"\toprule")
    L(r"\textbf{Entity} & \textbf{Cluster (MFA)} & \textbf{Silhouette} \\")
    L(r"\midrule")
    L(r"\endhead")
    for _, row in cl4.iterrows():
        entity = tex_escape(str(row["entity"]))
        cl = str(int(row["cluster_mfa"]))
        sil = fmt(row["silhouette_mfa"], 4)
        L(f"{entity} & {cl} & {sil} \\\\")
    L(r"\bottomrule")
    L(r"\multicolumn{3}{l}{\footnotesize Этап IV: MFA + K-Means. 49 объектов (включая Russia\_1, Russia\_2).} \\")
    L(r"\end{longtable}")

    L(r"")
    L(r"\end{document}")

    OUT_TEX.write_text("\n".join(lines), encoding="utf-8")
    print(f"LaTeX saved: {OUT_TEX}")


# =====================================================================
if __name__ == "__main__":
    build_excel()
    build_latex()
    print("Done.")
